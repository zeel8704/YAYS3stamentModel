import openpyxl
wb = openpyxl.load_workbook('d:/FinancialData/3-Statement Integrated Financial Model.xlsx', data_only=False)
ws = wb['Income Statement']
formats = set()
for r in range(1, 40):
    for c in range(1, 10):
        fmt = ws.cell(row=r, column=c).number_format
        if fmt != 'General':
            formats.add(fmt)

ws_ass = wb['Assumptions']
for r in range(1, 40):
    for c in range(1, 10):
        fmt = ws_ass.cell(row=r, column=c).number_format
        if fmt != 'General':
            formats.add(fmt)

print("Unique formats used:")
for f in formats:
    print(repr(f))
