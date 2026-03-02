import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class FinancialModel:
    def __init__(self, custom_assumptions=None):
        # Default Assumptions exactly as in reference
        self.assumptions = custom_assumptions or {
            "REVENUE ASSUMPTIONS": {
                "Year 1 Revenue (INR)": 5400000,
                "Revenue Growth Rate - Y2": 0.50,
                "Revenue Growth Rate - Y3": 0.35,
                "Revenue Growth Rate - Y4": 0.25,
                "Revenue Growth Rate - Y5": 0.15,
                "Gross Margin Change": 0.07,
                "Gross Margin %": 0.60,
            },
            "SG&A ASSUMPTIONS": {
                "Sales & Marketing (% Revenue)": 0.20,
                "General & Admin (% Revenue)": 0.10,
                "R&D (% Revenue)": 0.08,
            },
            "CAPEX & DEPRECIATION": {
                "Capex (% Revenue)": 0.05,
                "Useful Life (years)": 5.0,
                "Beginning PP&E, Net (INR)": 1400000,
            },
            "DEBT SCHEDULE": {
                "Beginning Debt (INR)": 600000,
                "Annual Interest Rate": 0.07,
                "Annual Debt Repayment (INR)": 100000,
            },
            "OTHER": {
                "Tax Rate": 0.30,
                "Beginning Cash (INR)": 800000,
                "Beginning Equity (INR)": 2000000,
            }
        }
        self.years = 5

class ExcelWriter:
    def __init__(self, model: FinancialModel):
        self.model = model
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active) # Remove default sheet
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.bold_font = Font(bold=True)
        self.italic_font = Font(italic=True)
        self.input_font = Font(color="0000FF", bold=False)  # Blue for hardcoded inputs
        self.formula_font = Font(color="000000", bold=False) # Black for formulas
        self.border_bottom = Border(bottom=Side(style='thin'))
        self.border_top = Border(top=Side(style='thin'))
        self.border_top_bottom = Border(top=Side(style='thin'), bottom=Side(style='double'))
        
        # Standard formats based on reference model
        self.fmt_number = '#,##0;(#,##0);"-"'
        self.fmt_percent = '0.0%;(0.0%);"-"'
        
    def _apply_header_style(self, cell):
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal="center")

    def _apply_sheet_beautification(self, ws, freeze='C3', zoom=85):
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = zoom
        if freeze:
            ws.freeze_panes = freeze

    def create_model(self, file_path_or_buffer):
        self._write_income_statement()
        self._write_balance_sheet()
        self._write_cash_flow_statement()
        self._write_capex_schedule()
        self._write_debt_schedule()
        self._protect_workbook()
        self.wb.save(file_path_or_buffer)

    def _protect_workbook(self):
        for ws in self.wb.worksheets:
            ws.protection.sheet = True
            ws.protection.password = 'financial_model_lock_99'
            ws.protection.enable()

    def _apply_standard_headers(self, ws, title):
        ws.column_dimensions['B'].width = 40
        ws.cell(row=1, column=2, value=title).font = self.bold_font
        ws.cell(row=2, column=2, value="Metric").font = self.bold_font
        ws.cell(row=2, column=2).fill = self.header_fill
        ws.cell(row=2, column=2).font = self.header_font
        
        for i in range(1, self.model.years + 1):
            col = 2 + i
            cell = ws.cell(row=2, column=col, value=f"Year {i}")
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_income_statement(self):
        ws = self.wb.create_sheet("Income Statement")
        self._apply_sheet_beautification(ws, freeze='C3')
        self._apply_standard_headers(ws, "Income Statement (INR)")
        
        roles = [
            (3, "Revenue", True, 0),
            (4, "Cost of Goods Sold", False, 1),
            (5, "Gross Profit", True, 0),
            (6, "Gross Margin %", False, 1),
            (8, "Operating Expenses", True, 0),
            (9, "Sales & Marketing", False, 1),
            (10, "General & Administrative", False, 1),
            (11, "Research & Development", False, 1),
            (12, "Depreciation", False, 1),
            (13, "Total Operating Expenses", True, 0),
            (15, "EBITDA", True, 0),
            (16, "EBITDA Margin", False, 1),
            (18, "EBIT (Operating Income)", True, 0),
            (19, "Operating Margin", False, 1),
            (20, "Interest Expense", False, 1),
            (22, "EBT (Pre-Tax Income)", True, 0),
            (23, "Income Tax Expense", False, 1),
            (25, "Net Income", True, 0),
            (26, "Net Margin %", False, 1)
        ]
        
        for r, name, is_bold, indent in roles:
            cell = ws.cell(row=r, column=2, value=name)
            if is_bold: cell.font = self.bold_font
            if indent: cell.alignment = Alignment(indent=indent)

        rev = self.model.assumptions["REVENUE ASSUMPTIONS"]
        sga = self.model.assumptions["SG&A ASSUMPTIONS"]
        other = self.model.assumptions["OTHER"]
        
        for i in range(1, self.model.years + 1):
            col = get_column_letter(2 + i)
            prev_col = get_column_letter(2 + i - 1) if i > 1 else None
            
            # Row 3: Revenue
            if i == 1:
                ws.cell(row=3, column=2+i, value=rev["Year 1 Revenue (INR)"])
            else:
                growth_rates = [0, 0, rev["Revenue Growth Rate - Y2"], rev["Revenue Growth Rate - Y3"], rev["Revenue Growth Rate - Y4"], rev["Revenue Growth Rate - Y5"]]
                ws.cell(row=3, column=2+i, value=f"={prev_col}3*(1+{growth_rates[i]})")
                
            # Row 4: COGS
            ws.cell(row=4, column=2+i, value=f"=-{col}3*(1-{col}6)")
            
            # Row 5: Gross Profit
            ws.cell(row=5, column=2+i, value=f"={col}3+{col}4").font = self.bold_font
            
            # Row 6: Gross Margin %
            if i == 1:
                ws.cell(row=6, column=2+i, value=rev["Gross Margin %"])
            else:
                ws.cell(row=6, column=2+i, value=f"={prev_col}6*(1+{rev['Gross Margin Change']})")
                
            # Rows 9,10,11
            ws.cell(row=9, column=2+i, value=f"=-{col}3*{sga['Sales & Marketing (% Revenue)']}")
            ws.cell(row=10, column=2+i, value=f"=-{col}3*{sga['General & Admin (% Revenue)']}")
            ws.cell(row=11, column=2+i, value=f"=-{col}3*{sga['R&D (% Revenue)']}")
            
            # Row 12: Depreciation
            ws.cell(row=12, column=2+i, value=f"=-'Balance Sheet'!{col}24")
            
            # Row 13: Total OpEx
            ws.cell(row=13, column=2+i, value=f"=SUM({col}9:{col}12)").font = self.bold_font
            
            # Row 15: EBITDA
            ws.cell(row=15, column=2+i, value=f"={col}18-{col}12").font = self.bold_font
            
            # Row 16: EBITDA Margin
            ws.cell(row=16, column=2+i, value=f"={col}15/{col}3")
            
            # Row 18: EBIT
            ws.cell(row=18, column=2+i, value=f"={col}5+{col}13").font = self.bold_font
            
            # Row 19: Operating Margin
            ws.cell(row=19, column=2+i, value=f"={col}18/{col}3")
            
            # Row 20: Interest Expense
            ws.cell(row=20, column=2+i, value=f"=-'Debt Schedule'!{col}6")
            
            # Row 22: EBT
            ws.cell(row=22, column=2+i, value=f"={col}18+{col}20").font = self.bold_font
            
            # Row 23: Income Tax Expense
            ws.cell(row=23, column=2+i, value=f"=IF({col}22>0,-{col}22*{other['Tax Rate']},0)")
            
            # Row 25: Net Income
            ws.cell(row=25, column=2+i, value=f"={col}22+{col}23").font = self.bold_font
            ws.cell(row=25, column=2+i).border = self.border_top_bottom
            
            # Row 26: Net Margin %
            ws.cell(row=26, column=2+i, value=f"=IF({col}3=0,0,{col}25/{col}3)")

            # Format percentages
            for r in [6, 16, 19, 26]:
                ws.cell(row=r, column=2+i).number_format = self.fmt_percent
            # Format numbers
            for r in [3, 4, 5, 9, 10, 11, 12, 13, 15, 18, 20, 22, 23, 25]:
                ws.cell(row=r, column=2+i).number_format = self.fmt_number

    def _write_balance_sheet(self):
        ws = self.wb.create_sheet("Balance Sheet")
        self._apply_sheet_beautification(ws, freeze='C3')
        self._apply_standard_headers(ws, "BALANCE SHEET (INR)")
        
        roles = [
            (4, "ASSETS", True, 0),
            (5, "Cash & Equivalents", False, 1),
            (6, "Accounts Receivable", False, 1),
            (7, "Inventory", False, 1),
            (8, "Total Current Assets", True, 0),
            (9, "PP&E, Gross", False, 1),
            (10, "Accumulated Depreciation", False, 1),
            (11, "PP&E, Net", True, 0),  # Not bold in source, wait, source says BOLD for C11 "FORMULA[=C9+C10] (BOLD)"
            (13, "Total Assets", True, 0),
            (15, "LIABILITIES & EQUITY", True, 0),
            (16, "Accounts Payable", False, 1),
            (17, "Accrued Liabilities", False, 1),
            (18, "Total Current Liabilities", True, 0),
            (19, "Long-Term Debt", False, 1),
            (20, "Total Liabilities", True, 0),
            (21, "Common Equity / Paid-In Capital", False, 1),
            (22, "Retained Earnings", False, 1),
            (23, "Total Stockholders' Equity", True, 0),
            (24, "Depreciation Expense (plug for IS)", False, 1),
            (26, "TOTAL LIABILITIES & EQUITY", True, 0),
            (28, "Balance Sheet Check", True, 0)
        ]
        
        for r, name, is_bold, indent in roles:
            cell = ws.cell(row=r, column=2, value=name)
            if is_bold: cell.font = self.bold_font
            if r == 11:
                pass # C11 is bold, B11 is indented
                cell.font = Font(bold=False)
                cell.alignment = Alignment(indent=1)
            if indent and r != 11: cell.alignment = Alignment(indent=indent)

        for i in range(1, self.model.years + 1):
            col = get_column_letter(2 + i)
            prev_col = get_column_letter(2 + i - 1) if i > 1 else None
            
            ws.cell(row=5, column=2+i, value=f"={col}26-{col}6-{col}7-{col}11")
            ws.cell(row=6, column=2+i, value=f"='Income Statement'!{col}3*0.1")
            ws.cell(row=7, column=2+i, value=f"='Income Statement'!{col}3*0.08")
            ws.cell(row=8, column=2+i, value=f"={col}5+{col}6+{col}7").font = self.bold_font
            
            ws.cell(row=9, column=2+i, value=f"='Capex Schedule'!{col}6")
            ws.cell(row=10, column=2+i, value=f"=-'Capex Schedule'!{col}9")
            ws.cell(row=11, column=2+i, value=f"={col}9+{col}10").font = self.bold_font
            
            ws.cell(row=13, column=2+i, value=f"={col}8+{col}11").font = self.bold_font
            ws.cell(row=13, column=2+i).border = self.border_top_bottom
            
            ws.cell(row=16, column=2+i, value=f"='Income Statement'!{col}3*0.05")
            ws.cell(row=17, column=2+i, value=f"='Income Statement'!{col}3*0.03")
            ws.cell(row=18, column=2+i, value=f"={col}16+{col}17").font = self.bold_font
            
            ws.cell(row=19, column=2+i, value=f"='Debt Schedule'!{col}7")
            ws.cell(row=20, column=2+i, value=f"={col}18+{col}19").font = self.bold_font
            
            other = self.model.assumptions["OTHER"]
            if i == 1:
                ws.cell(row=21, column=2+i, value=other["Beginning Equity (INR)"])
            else:
                ws.cell(row=21, column=2+i, value=f"={prev_col}21")
                
            if i == 1:
                ws.cell(row=22, column=2+i, value=f"='Income Statement'!C25")
            else:
                ws.cell(row=22, column=2+i, value=f"={prev_col}22+'Income Statement'!{col}25")
                
            ws.cell(row=23, column=2+i, value=f"={col}21+{col}22").font = self.bold_font
            ws.cell(row=24, column=2+i, value=f"='Capex Schedule'!{col}8")
            
            ws.cell(row=26, column=2+i, value=f"={col}20+{col}23").font = self.bold_font
            ws.cell(row=26, column=2+i).border = self.border_top_bottom
            
            ws.cell(row=28, column=2+i, value=f"={col}13-{col}26")
            
            for r in [5,6,7,8,9,10,11,13,16,17,18,19,20,21,22,23,24,26,28]:
                ws.cell(row=r, column=2+i).number_format = self.fmt_number

    def _write_cash_flow_statement(self):
        ws = self.wb.create_sheet("Cash Flow Statement")
        self._apply_sheet_beautification(ws, freeze='C3')
        self._apply_standard_headers(ws, "CASH FLOW STATEMENT (INR)")
        
        roles = [
            (4, "Operating Activities", True, 0),
            (5, "Net Income", False, 1),
            (6, "Add: Depreciation & Amortization", False, 1),
            (7, "Change in Accounts Receivable", False, 1),
            (8, "Change in Inventory", False, 1),
            (9, "Change in Accounts Payable", False, 1),
            (10, "Change in Accrued Liabilities", False, 1),
            (12, "Net Cash from Operations", True, 0),
            (14, "Investing Activities", True, 0),
            (15, "Capital Expenditures", False, 1),
            (17, "Net Cash from Investing", True, 0),
            (19, "Financing Activities", True, 0),
            (20, "Debt Repayment", False, 1),
            (22, "Net Cash from Financing", True, 0),
            (24, "Beginning Cash", False, 0),
            (25, "Ending Cash", False, 0),
            (27, "Net Change in Cash", True, 0)
        ]
        
        for r, name, is_bold, indent in roles:
            cell = ws.cell(row=r, column=2, value=name)
            if is_bold: cell.font = self.bold_font
            if indent: cell.alignment = Alignment(indent=indent)

        for i in range(1, self.model.years + 1):
            col = get_column_letter(2 + i)
            prev_col = get_column_letter(2 + i - 1) if i > 1 else None
            
            ws.cell(row=5, column=2+i, value=f"='Income Statement'!{col}25")
            ws.cell(row=6, column=2+i, value=f"='Capex Schedule'!{col}8")
            
            if i == 1:
                ws.cell(row=7, column=2+i, value=f"=-'Balance Sheet'!C6")
                ws.cell(row=8, column=2+i, value=f"=-'Balance Sheet'!C7")
                ws.cell(row=9, column=2+i, value=f"='Balance Sheet'!C16")
                ws.cell(row=10, column=2+i, value=f"='Balance Sheet'!C17")
            else:
                ws.cell(row=7, column=2+i, value=f"=-('Balance Sheet'!{col}6-'Balance Sheet'!{prev_col}6)")
                ws.cell(row=8, column=2+i, value=f"=-('Balance Sheet'!{col}7-'Balance Sheet'!{prev_col}7)")
                ws.cell(row=9, column=2+i, value=f"='Balance Sheet'!{col}16-'Balance Sheet'!{prev_col}16")
                ws.cell(row=10, column=2+i, value=f"='Balance Sheet'!{col}17-'Balance Sheet'!{prev_col}17")
                
            ws.cell(row=12, column=2+i, value=f"=SUM({col}5:{col}10)").font = self.bold_font
            ws.cell(row=12, column=2+i).border = self.border_top_bottom
            
            ws.cell(row=15, column=2+i, value=f"=-'Capex Schedule'!{col}5")
            
            ws.cell(row=17, column=2+i, value=f"={col}15").font = self.bold_font
            ws.cell(row=17, column=2+i).border = self.border_top_bottom
            
            ws.cell(row=20, column=2+i, value=f"='Debt Schedule'!{col}5")
            
            ws.cell(row=22, column=2+i, value=f"={col}20").font = self.bold_font
            ws.cell(row=22, column=2+i).border = self.border_top_bottom
            
            other = self.model.assumptions["OTHER"]
            if i == 1:
                ws.cell(row=24, column=2+i, value=other["Beginning Cash (INR)"])
            else:
                ws.cell(row=24, column=2+i, value=f"={prev_col}25")
                
            ws.cell(row=25, column=2+i, value=f"={col}24+{col}27")
            
            ws.cell(row=27, column=2+i, value=f"={col}12+{col}17+{col}22").font = self.bold_font
            ws.cell(row=27, column=2+i).border = self.border_top_bottom
            
            for r in [5,6,7,8,9,10,12,15,17,20,22,24,25,27]:
                ws.cell(row=r, column=2+i).number_format = self.fmt_number

    def _write_capex_schedule(self):
        ws = self.wb.create_sheet("Capex Schedule")
        self._apply_sheet_beautification(ws, freeze='C3')
        self._apply_standard_headers(ws, "CAPEX & DEPRECIATION SCHEDULE (INR)")
        
        roles = [
            (4, "Beginning PP&E, Gross", False, 0),
            (5, "Capital Expenditures (Additions)", False, 0),
            (6, "Ending PP&E, Gross", True, 0),
            (7, "Accumulated Depreciation (Beg)", False, 0),
            (8, "Depreciation Expense", True, 0),
            (9, "Accumulated Depreciation (End)", False, 0),
            (10, "PP&E, Net (Ending)", True, 0),
        ]
        
        for r, name, is_bold, indent in roles:
            cell = ws.cell(row=r, column=2, value=name)
            if is_bold: cell.font = self.bold_font
            if indent: cell.alignment = Alignment(indent=indent)

        for i in range(1, self.model.years + 1):
            col = get_column_letter(2 + i)
            prev_col = get_column_letter(2 + i - 1) if i > 1 else None
            
            if i == 1:
                ws.cell(row=4, column=2+i, value=1500000)
            else:
                ws.cell(row=4, column=2+i, value=f"={prev_col}6")
                
            cap = self.model.assumptions["CAPEX & DEPRECIATION"]
            ws.cell(row=5, column=2+i, value=f"='Income Statement'!{col}3*{cap['Capex (% Revenue)']}")
            ws.cell(row=6, column=2+i, value=f"={col}4+{col}5").font = self.bold_font
            
            if i == 1:
                ws.cell(row=7, column=2+i, value=0)
            else:
                ws.cell(row=7, column=2+i, value=f"={prev_col}9")
                
            ws.cell(row=8, column=2+i, value=f"={col}6/{cap['Useful Life (years)']}").font = self.bold_font
            ws.cell(row=9, column=2+i, value=f"={col}7+{col}8")
            ws.cell(row=10, column=2+i, value=f"={col}6-{col}9").font = self.bold_font
            ws.cell(row=10, column=2+i).border = self.border_top_bottom
            
            for r in [4,5,6,7,8,9,10]:
                ws.cell(row=r, column=2+i).number_format = self.fmt_number

    def _write_debt_schedule(self):
        ws = self.wb.create_sheet("Debt Schedule")
        self._apply_sheet_beautification(ws, freeze='C3')
        self._apply_standard_headers(ws, "DEBT SCHEDULE (INR)")
        
        roles = [
            (4, "Beginning Debt Balance", False, 1),
            (5, "Debt Repayment", False, 1),
            (6, "Interest Expense", True, 1),
            (7, "Ending Debt Balance", True, 1),
        ]
        
        for r, name, is_bold, indent in roles:
            cell = ws.cell(row=r, column=2, value=name)
            if is_bold: cell.font = self.bold_font
            if indent: cell.alignment = Alignment(indent=indent)

        for i in range(1, self.model.years + 1):
            col = get_column_letter(2 + i)
            prev_col = get_column_letter(2 + i - 1) if i > 1 else None
            
            debt = self.model.assumptions["DEBT SCHEDULE"]
            if i == 1:
                ws.cell(row=4, column=2+i, value=debt["Beginning Debt (INR)"])
            else:
                ws.cell(row=4, column=2+i, value=f"={prev_col}7")
                
            ws.cell(row=5, column=2+i, value=f"=-MIN({debt['Annual Debt Repayment (INR)']},{col}4)")
            ws.cell(row=6, column=2+i, value=f"={col}4*{debt['Annual Interest Rate']}").font = self.bold_font
            ws.cell(row=7, column=2+i, value=f"={col}4+{col}5").font = self.bold_font
            ws.cell(row=7, column=2+i).border = self.border_top_bottom
            
            for r in [4,5,6,7]:
                ws.cell(row=r, column=2+i).number_format = self.fmt_number


