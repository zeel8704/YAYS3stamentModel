import openpyxl
wb = openpyxl.load_workbook('d:/FinancialData/3-Statement Integrated Financial Model.xlsx', data_only=False)

for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    print(f"\nSheet: {sheetname}")
    try:
        print(f"show_gridlines: {ws.sheet_view.showGridLines}")
        print(f"zoom_scale: {ws.sheet_view.zoomScale}")
        print(f"freeze_panes: {ws.freeze_panes}")
    except Exception as e:
        print(f"Error reading view properties: {e}")
    # font of A1, A2
    font_a1 = ws['B2'].font
    font_a2 = ws['C3'].font
    if font_a2:
        print(f"Font in C3: {font_a2.name}, size: {font_a2.sz}, color: {font_a2.color.rgb if font_a2.color and hasattr(font_a2.color, 'rgb') else 'None'}")
    
    col_width_B = ws.column_dimensions['B'].width
    col_width_C = ws.column_dimensions['C'].width
    print(f"Col B width: {col_width_B}")
    print(f"Col C width: {col_width_C}")
